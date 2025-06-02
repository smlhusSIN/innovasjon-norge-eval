import openai
import pandas as pd
from typing import List, Dict, Tuple
import os
from dotenv import load_dotenv
import PyPDF2
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import re
from openai import OpenAIError, APITimeoutError, APIConnectionError, AuthenticationError, BadRequestError, RateLimitError
from openpyxl.chart import RadarChart, Reference
from evaluate_nic_application import evaluate_nic_application, create_nic_excel_report

# Load environment variables from .env file
load_dotenv()

# Set up OpenAI API key
openai_api_key = os.getenv("OPENAI_API_KEY")
client = openai.OpenAI(api_key=openai_api_key)

# Evaluation questions organized by category
EVALUATION_QUESTIONS = {
    "Problemløsning og marked": [
        "Hvor godt er problemet/behovet/dagens situasjon beskrevet, inkludert dagens alternative løsninger?",
        "Hvor godt er løsningsbeskrivelsen beskrevet og er den tydelig avgrenset?",
        "Hvor godt er FoU-utfordringer knyttet til løsningen beskrevet?",
        "Hvor godt er det forklart hvorfor FoU-utfordringer ikke er løst tidligere?",
        "Hvor godt er dagens konkurrenter og konkurransebilde beskrevet?",
        "Hvor godt beskrevet er forskjellene og forbedringene fra dagens alternativer?",
        "Hvor tydelig er de unike aspektene ved løsningen beskrevet (det må være vesentlige forskjeller fra dagens løsninger)?"
    ],
    "Kapning": [
        "Hvor godt er den kortsiktige effekten av prosjektresultatet for selskapet beskrevet? (kunder, inntekt, arbeidsplasser o.l.)",
        "Hvor godt er den langsiktige effekten av prosjektresultatet for selskapet beskrevet? (kunder, inntekt, arbeidsplasser o.l.)",
        "Hvor tydelig er det at prosjektet bygger opp under selskapets langsiktige ambisjon og strategi?",
        "Hvor godt er verdiskapningen for samfunnet beskrevet?",
        "Hvor godt er miljøgevinsten beskrevet?",
        "Hvor godt er gevinster som bærekraft, likestilling og mangfold beskrevet?",
        "Hvor toverig fremstår markedsbeskrivelsen, i både nasjonal og internasjonal perspektiv?",
        "Hvor tydelig og realistisk er forretningsmodellen beskrevet?",
        "Hvor godt er kundeverdien beskrevet?",
        "Hvor godt er interaksjon med kunder beskrevet som en del av prosjektet?"
    ],
    "Gjennomføringsevne": [
        "Hvor godt er selskapets nåværende situasjon, målsetting og strategi beskrevet?",
        "Hvor tydelig og avgrenset er prosjektet beskrevet?",
        "Hvor godt er aktivitetene/arbeidspakkene i prosjektet beskrevet?",
        "Hvor godt er roller og ansvar beskrevet?",
        "For FoU-prosjekter: hvor godt er håndteringen av disse beskrevet?",
        "Hvor godt er teamets og selskapets erfaring og kompetanse beskrevet?",
        "Hvis dere har partnere eller kunder som er en del av prosjektet, hvor godt beskrevet er disse inkludert bidrag under og etter prosjektet?",
        "Hvor gode forutsetninger har selskapet for å kunne realisere 'go-to-market'-planer? f.eks. etablering av kanaler til markedet?",
        "Hvis dere har leverandører, hvor godt er disse og deres fortinn beskrevet?",
        "Hvor godt er finansieringsplanen under og etter prosjektet beskrevet i tekst og vedlegg? Er dere tydelig likviditet i perioden?",
        "Hvor godt kommer det frem at selskapet har spesifikke fordeler som gjør selskapet spesielt egnet til å forvalte investeringen?"
    ],
    "Statsstøtte-effekt av støtte fra Innovasjon Norge": [
        "Hvor godt kommer det frem at dere er avhengig av støtte/lån for å realisere/akselerere prosjektet?",
        "Hvor godt er teknisk risiko beskrevet i søknaden?",
        "Hvor godt er markeds-/kommersiell risiko beskrevet i søknaden?",
        "Hvor godt er klimarisiko (negativ effekt av prosjekt-antagelser og klimaendringer) beskrevet?",
        "Totalt sett, er risikoen i prosjektet stor nok til at det aktuelt for en bank eller investor å investere i prosjektet uten statsstøtte?",
        "Hvor godt beskrevet er scenarioene med og uten støtte fra Innovasjon Norge for selskapet?",
        "Hvor godt er et potensielle investorer beskrevet, helst navngitt, gitt støtte fra Innovasjon Norge? Beskrevet i tekst eller vedlegg?",
        "Hvor godt kommer det frem at dere har kunder som vil kjøpe, gitt gjennomføring av prosjekt (LOI eller tilsvarende)?",
        "Hvor godt er markedssituasjonen beskrevet, og hvordan støtte fra IN kan posisjonere dere?"
    ],
    "Gjennomføring og detaljer": [
        "Svarer søknaden på alle krav som etterspørres i IN sin søknadsportal?",
        "Er språket korrekturlest, både med tanke på tegnsetting og rettskrivning? (Nei = 0, Ja =3)",
        "Hvor overbevisende/tillitsvekkende er språket i søknaden?",
        "I hvor stor grad er språket kort og konsist?",
        "I hvor stor grad er 'buzzwords' unngått?",
        "I hvor stor grad har det blitt benyttet kildehenvisninger for dataunderlag og argumenter?",
        "Er timesatser for de forskjellige personalkategoriene innenfor grensene til IN (Nei = 0, Ja =3)",
        "Er maksimal støttegrad for de forskjellige aktivitetene tilpasset deres bedrift? (Nei = 0, Ja =3)",
        "Hvis det søkes om lån, hvor godt er sikkerheten for lånet beskrevet og dokumentert?",
        "Er alle tall dobbeltsjekket opp mot prosjektoppsettet deres?"
    ],
    "Krav fra IN": [
        "Krav fra IN: Siste års regnskap",
        "Krav fra IN: Perioderegnskab, ikke eldre enn 3 mnd",
        "Krav fra IN: Driftsbudsjett for bedriftens virksomhet de neste 3 årene",
        "Krav fra IN: Eventuelle budsjetter/lønnsomhetsberegninger",
        "Selskapsrepresentasjon",
        "Prosjektpresentasjon",
        "Organisasjonskart inkludert CV på nøkkelpersoner i prosjektet",
        "Kundeavtaler, Letter of Intent (LOI), eller annet som understreker markedsbehovet",
        "Termsheet, Intensjonsbrev eller lignende som dokumenterer at dere har investor som vil investere gitt støtte fra IN.",
        "For vedlegg utarbeidet i Excel-format, er disse sendt i originalformatet?"
    ]
}

# Spørsmål for oppstart 1 (fra bildet)
EVALUATION_QUESTIONS_OPPSTART_1 = {
    "Problemløsning og marked": [
        "Hvor godt er problemet/behovet/dagens situasjon beskrevet, inkludert dagens alternative løsninger?",
        "Hvor godt er løsningsbeskrivelsen beskrevet og er den tydelig avgrenset?",
        "Hvor godt er dagens konkurrenter og konkurransebilde beskrevet?",
        "Hvor godt er forskjellene og forbedringene fra dagens alternativer beskrevet?",
        "Hvor tydelig er de unike aspektene ved løsningen beskrevet (det må være VESENTLIGE forskjeller fra dagens løsninger)?"
    ],
    "Verdiskapning": [
        "Hvor godt er det kommersielle potensialet for selskapet beskrevet? (kunder, inntekt, arbeidsplasser o.l.)",
        "Hvor godt er verdiskapningen for samfunnet beskrevet?",
        "Hvor er gevinster som miljø, bærekraft, likestilling og mangfold beskrevet?",
        "Hvor tydelig og realistisk er forretningsmodellen beskrevet?",
        "Hvor godt er kundeverdien beskrevet?",
        "Hvor godt er kundegruppe og markedet beskrevet?"
    ],
    "Gjennomføringsevne": [
        "Hvor tydelig og avgrenset er prosjektet beskrevet?",
        "Hvor godt er målsettingen for prosjektet tilpasset hensikten med ordningen -> validering av problem, løsning og marked?",
        "Hvor godt er suksesskriterier for å kunne gå videre med forretningsideen etter endt prosjekt beskrevet?",
        "Hvor godt er teamets og selskapets relevante erfaring og relevant kompetanse beskrevet?",
        "Hvor er investorer, inkubatorer, rådgivere eller andre støttespillere beskrevet?",
        "Hvor godt er kundegruppe og markedet beskrevet?"
    ],
    "Utløsende effekt av støtte fra Innovasjon Norge": [
        "Hvor godt kommer det frem at dere er avhengig av tilskudd for å realisere prosjektet?",
        "Hvor godt er risikoen prosjektet skal redusere beskrevet i søknaden?",
        "Hvis dere har investorer som ønsker å investere i dere, i etterkant av prosjektet, er dette beskrevet?"
    ],
    "Søknadsutforming og detaljer": [
        "Svarer søknaden på alle krav som etterspørres i IN sin søknadsportal?",
        "Er språket korrekturlest, både med tanke på tegnsetting og rettskrivning? (Nei = 0, Ja =3)",
        "Hvor overbevisende/tillitsvekkende er språket i søknaden?",
        "I hvor stor grad er språket kort og konsist?",
        "I hvor stor grad er 'buzzwords' unngått?",
        "I hvor stor grad har det blitt benyttet kildehenvisninger for dataunderlag og argumenter?"
    ],
    "Vedlegg (Nei = 0, Ja = 3)": [
        "Finansiell modell eller likviditetsbudsjett for selskapet",
        "Selskaspresentasjon / Pitch-deck",
        "Prosjektpresentasjon",
        "Forretningsmodell (hvis ikke en del av selskapspresentasjonen), som 'lean business canvas' eller tilsvarende",
        "Konkurrentanalyse (hvis ikke en del av prosjektpresentasjon)",
        "Siste til løsningsforslag",
        "For vedlegg utarbeidet i Excel-format, er disse sendt i originalformatet?"
    ]
}

def read_application_text(filename: str = None) -> tuple[str, str]:
    """Read the application text from a PDF file."""
    if filename is None:
        # Find PDF files in current directory
        pdf_files = glob.glob("*.pdf")
        if not pdf_files:
            raise FileNotFoundError("❌ FEIL: Ingen PDF-filer funnet i mappen. Legg til en PDF-fil med søknaden.")
        
        if len(pdf_files) == 1:
            filename = pdf_files[0]
            print(f"📄 Bruker PDF-fil: {filename}")
        else:
            # Multiple PDF files found, let user choose
            print(f"\n📁 Fant {len(pdf_files)} PDF-filer i mappen:")
            for i, pdf_file in enumerate(pdf_files, 1):
                print(f"  {i}. {pdf_file}")
            
            while True:
                try:
                    choice = input(f"\nVelg hvilken PDF du vil evaluere (1-{len(pdf_files)}): ")
                    choice_num = int(choice)
                    if 1 <= choice_num <= len(pdf_files):
                        filename = pdf_files[choice_num - 1]
                        print(f"📄 Valgt PDF-fil: {filename}")
                        break
                    else:
                        print(f"❌ Ugyldig valg. Velg et tall mellom 1 og {len(pdf_files)}.")
                except ValueError:
                    print("❌ Ugyldig input. Skriv inn et tall.")
                except KeyboardInterrupt:
                    print("\n🛑 Avbrutt av bruker.")
                    raise
    
    try:
        text = ""
        with open(filename, 'rb') as file:
            try:
                pdf_reader = PyPDF2.PdfReader(file)
                total_pages = len(pdf_reader.pages)
                print(f"📖 Leser {total_pages} sider fra PDF...")
                
                for i, page in enumerate(pdf_reader.pages, 1):
                    page_text = page.extract_text()
                    text += page_text + "\n"
                    if i % 5 == 0:  # Show progress every 5 pages
                        print(f"   📄 Behandlet side {i}/{total_pages}")
                
            except PyPDF2.errors.PdfReadError:
                raise Exception(f"❌ FEIL: Kunne ikke lese PDF-filen '{filename}'. Filen kan være korrupt eller passordbeskyttet.")
            except Exception as e:
                raise Exception(f"❌ FEIL: Problem ved lesing av PDF-innhold: {e}")
        
        if not text.strip():
            raise ValueError(f"❌ FEIL: Ingen tekst kunne ekstraheres fra PDF-filen '{filename}'. Filen kan være tom eller inneholde kun bilder.")
        
        print(f"✅ Tekst ekstrahert fra {total_pages} sider")
        return text, filename
        
    except FileNotFoundError:
        raise FileNotFoundError(f"❌ FEIL: PDF-filen '{filename}' ble ikke funnet. Sjekk at filen eksisterer i mappen.")
    except PermissionError:
        raise Exception(f"❌ FEIL: Ingen tilgang til å lese filen '{filename}'. Sjekk filtillatelser.")
    except Exception as e:
        if "❌ FEIL:" in str(e):
            raise  # Re-raise our custom errors as-is
        else:
            raise Exception(f"❌ FEIL: Uventet problem ved lesing av PDF: {e}")

def get_score_from_openai(question: str, application_text: str) -> Tuple[int, str]:
    """Get score and comment from OpenAI API for a specific question."""
    prompt = f"""Basert på følgende søknad, gi en score fra 0-3 for dette spørsmålet: {question}
    
    Søknad: {application_text}
    
    Svar i følgende format:
    Score: [0-3]
    Kommentar: [kort kommentar]"""
    
    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": "Du er en ekspert på å evaluere søknader til Innovasjon Norge. Gi en score fra 0-3 og en kort kommentar."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3,
            max_tokens=200
        )
        
        # Parse the response
        response_text = response.choices[0].message.content
        
        # Find score and comment lines
        score_lines = [line for line in response_text.split('\n') if line.startswith('Score:')]
        comment_lines = [line for line in response_text.split('\n') if line.startswith('Kommentar:')]
        
        if not score_lines:
            raise ValueError(f"Kunne ikke finne 'Score:' i OpenAI-responsen: {response_text}")
        if not comment_lines:
            raise ValueError(f"Kunne ikke finne 'Kommentar:' i OpenAI-responsen: {response_text}")
        
        score_text = score_lines[0].split(':')[1].strip()
        comment = comment_lines[0].split(':')[1].strip()
        
        # Validate and convert score
        try:
            score = int(score_text)
            if score < 0 or score > 3:
                raise ValueError(f"Score må være mellom 0-3, fikk: {score}")
        except ValueError as ve:
            raise ValueError(f"Kunne ikke konvertere score til tall: '{score_text}'. {ve}")
        
        return score, comment
    
    except AuthenticationError as e:
        raise Exception(f"❌ FEIL: OpenAI API-nøkkel er ugyldig. Sjekk at OPENAI_API_KEY er riktig satt i .env filen. Detaljer: {e}")
    except RateLimitError as e:
        raise Exception(f"❌ FEIL: OpenAI API rate limit nådd. Vent litt og prøv igjen. Detaljer: {e}")
    except APIConnectionError as e:
        raise Exception(f"❌ FEIL: Kunne ikke koble til OpenAI API. Sjekk internettforbindelsen din. Detaljer: {e}")
    except BadRequestError as e:
        raise Exception(f"❌ FEIL: Ugyldig forespørsel til OpenAI API: {e}. Sjekk at input ikke er for lang eller inneholder ugyldige parametre.")
    except ValueError as ve:
        raise Exception(f"❌ FEIL: Problem med å tolke OpenAI-respons: {ve}")
    except OpenAIError as e:
        raise Exception(f"❌ FEIL: OpenAI-feil: {type(e).__name__}: {e}")
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        raise Exception(f"❌ FEIL: Uventet feil ved OpenAI API-kall: {type(e).__name__}: {e}\nTraceback:\n{tb}")

def evaluate_application(application_text: str, pdf_filename: str = None, evaluation_questions=None) -> pd.DataFrame:
    """Evaluate the application using OpenAI API and return results as DataFrame."""
    results = []
    
    if evaluation_questions is None:
        evaluation_questions = EVALUATION_QUESTIONS
    # Calculate total number of questions for progress tracking
    total_questions = sum(len(questions) for questions in evaluation_questions.values())
    current_question = 0
    
    for category, questions in evaluation_questions.items():
        print(f"\n📋 Evaluerer kategori: {category}")
        for question in questions:
            current_question += 1
            print(f"  ⏳ Spørsmål {current_question}/{total_questions}: {question[:50]}...")
            
            try:
                score, comment = get_score_from_openai(question, application_text)
                print(f"  ✅ Score: {score}/3")
                
                results.append({
                    "Kategori": category,
                    "Spørsmål": question,
                    "Score": score,
                    "Kommentar": comment
                })
            except Exception as e:
                print(f"  ❌ Feil ved evaluering av spørsmål: {e}")
                # Add a fallback entry with error information
                results.append({
                    "Kategori": category,
                    "Spørsmål": question,
                    "Score": 0,
                    "Kommentar": f"Feil ved evaluering: {str(e)[:100]}..."
                })
                # Ask user if they want to continue
                print(f"  ⚠️  Vil du fortsette med neste spørsmål? (Trykk Enter for å fortsette, Ctrl+C for å avbryte)")
                try:
                    input()
                except KeyboardInterrupt:
                    print("\n🛑 Evaluering avbrutt av bruker.")
                    raise
    
    print(f"\n🎉 Evaluering fullført! {total_questions} spørsmål behandlet.")
    return pd.DataFrame(results)

def create_excel_report(results_df: pd.DataFrame, pdf_filename: str, excel_filename: str, oppstartstype: str = "") -> None:
    """Create a formatted Excel report with summary and detailed results."""
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Søknadsevaluering"
    
    # Define styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    summary_font = Font(bold=True, size=12)
    summary_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    category_font = Font(bold=True, size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Calculate summary statistics
    summary = results_df.groupby('Kategori')['Score'].mean().round(2)
    total_score = summary.mean()
    
    # Determine overall assessment
    if total_score >= 2.5:
        assessment = "🎉 Utmerket søknad! Høy sannsynlighet for godkjenning."
        assessment_color = "C6EFCE"  # Light green
    elif total_score >= 2.0:
        assessment = "👍 God søknad med potensial. Noen forbedringer kan styrke den."
        assessment_color = "FFEB9C"  # Light yellow
    elif total_score >= 1.5:
        assessment = "⚠️ Søknaden trenger forbedringer i flere områder."
        assessment_color = "FFEB9C"  # Light yellow
    else:
        assessment = "🔴 Søknaden har betydelige svakheter som bør adresseres."
        assessment_color = "FFC7CE"  # Light red
    
    current_row = 1
    
    # Title
    ws.merge_cells(f'A{current_row}:D{current_row}')
    title_cell = ws[f'A{current_row}']
    if oppstartstype:
        title_cell.value = f"INNOVASJON NORGE - SØKNADSEVALUERING ({oppstartstype})"
    else:
        title_cell.value = "INNOVASJON NORGE - SØKNADSEVALUERING"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = center_alignment
    current_row += 2
    
    # PDF filename
    ws[f'A{current_row}'] = "Evaluert søknad:"
    ws[f'A{current_row}'].font = summary_font
    ws[f'B{current_row}'] = pdf_filename
    ws.merge_cells(f'B{current_row}:D{current_row}')
    current_row += 2
    
    # Overall score
    ws[f'A{current_row}'] = "TOTAL GJENNOMSNITTSSCORE:"
    ws[f'A{current_row}'].font = Font(bold=True, size=14)
    ws[f'B{current_row}'] = f"{total_score:.2f}/3.0"
    ws[f'B{current_row}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws[f'B{current_row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws[f'B{current_row}'].alignment = center_alignment
    current_row += 1
    
    # Assessment
    ws.merge_cells(f'A{current_row}:D{current_row}')
    assessment_cell = ws[f'A{current_row}']
    assessment_cell.value = assessment
    assessment_cell.font = Font(bold=True, size=12)
    assessment_cell.fill = PatternFill(start_color=assessment_color, end_color=assessment_color, fill_type="solid")
    assessment_cell.alignment = center_alignment
    current_row += 2
    
    # Category summary header
    ws[f'A{current_row}'] = "SAMMENDRAG PER KATEGORI"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws.merge_cells(f'A{current_row}:C{current_row}')
    current_row += 1
    
    # Category summary
    for kategori, score in summary.items():
        emoji = "🟢" if score >= 2.5 else "🟡" if score >= 1.5 else "🔴"
        ws[f'A{current_row}'] = f"{emoji} {kategori}"
        ws[f'B{current_row}'] = f"{score}/3.0"
        ws[f'C{current_row}'] = score  # Tallverdi for diagrammet
        ws[f'B{current_row}'].alignment = center_alignment
        # Color coding
        if score >= 2.5:
            fill_color = "C6EFCE"  # Green
        elif score >= 1.5:
            fill_color = "FFEB9C"  # Yellow
        else:
            fill_color = "FFC7CE"  # Red
        ws[f'B{current_row}'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        current_row += 1
    
    current_row += 2
    
    # Detailed results header
    ws[f'A{current_row}'] = "DETALJERTE RESULTATER"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws.merge_cells(f'A{current_row}:D{current_row}')
    current_row += 1
    
    # Column headers for detailed results
    headers = ['Kategori', 'Spørsmål', 'Score', 'Kommentar']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = category_font
        cell.fill = summary_fill
        cell.alignment = center_alignment
        cell.border = border
    current_row += 1
    
    # Add detailed results
    for _, row in results_df.iterrows():
        ws.cell(row=current_row, column=1, value=row['Kategori']).border = border
        ws.cell(row=current_row, column=2, value=row['Spørsmål']).border = border
        
        score_cell = ws.cell(row=current_row, column=3, value=f"{row['Score']}/3")
        score_cell.border = border
        score_cell.alignment = center_alignment
        
        # Color code scores
        if row['Score'] >= 2.5:
            score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif row['Score'] >= 1.5:
            score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        comment_cell = ws.cell(row=current_row, column=4, value=row['Kommentar'])
        comment_cell.border = border
        comment_cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        current_row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 80
    
    # Set row heights for better readability
    for row in range(1, current_row):
        if row > len(summary) + 10:  # Detailed results section
            ws.row_dimensions[row].height = 60
    
    # Save the workbook
    wb.save(excel_filename)

def main():
    print("🚀 Starter søknadsevaluering...")
    
    # Velg evalueringstype
    print("\nVelg evalueringstype:")
    print("  1. Innovasjon Norge - Oppstart 1")
    print("  2. Innovasjon Norge - Oppstart 2")
    print("  3. Innovasjon Norge - Oppstart 3")
    print("  4. NIC Klyngeevaluering")
    while True:
        valg = input("Skriv inn 1, 2, 3 eller 4 og trykk Enter: ").strip()
        if valg in {"1", "2", "3", "4"}:
            break
        print("Ugyldig valg. Prøv igjen.")
    
    if valg == "4":
        # NIC evaluering
        print("\n🚀 Starter NIC Klyngeevaluering...")
        
        # Check if API key is set
        if not openai_api_key:
            print("❌ FEIL: OPENAI_API_KEY ikke funnet.")
            print("💡 LØSNING: Opprett en .env fil i samme mappe med innholdet:")
            print("   OPENAI_API_KEY=din-api-nøkkel-her")
            print("📖 Du kan få en API-nøkkel fra: https://platform.openai.com/api-keys")
            return
        
        try:
            # Read application text
            print("\n🔍 Leser klyngesøknad fra PDF...")
            application_text, selected_pdf = read_application_text()
            print(f"✅ PDF lest inn - {len(application_text):,} tegn funnet")
            
            # Create unique filename based on PDF name
            pdf_base_name = selected_pdf.replace('.pdf', '').replace(' ', '_')
            # Remove special characters that might cause issues
            pdf_base_name = re.sub(r'[^\w\-_]', '', pdf_base_name)
            excel_filename = f"nic_evaluering_resultat_{pdf_base_name}.xlsx"
            
            # Validate text length
            if len(application_text) < 100:
                print("⚠️  ADVARSEL: Søknadsteksten virker veldig kort. Sjekk at PDF-en inneholder tekst.")
                response = input("Vil du fortsette likevel? (j/n): ")
                if response.lower() != 'j':
                    print("🛑 Evaluering avbrutt.")
                    return
            
            # Evaluate application
            print("\n🤖 Starter AI-evaluering av klyngesøknaden...")
            print("Dette kan ta noen minutter avhengig av søknadens lengde.")
            print("💡 Tips: Du kan avbryte med Ctrl+C hvis nødvendig.")
            
            results_df = evaluate_nic_application(application_text, selected_pdf)
            
            # Create Excel report
            print(f"\n📊 Lager formatert Excel-rapport: {excel_filename}")
            try:
                create_nic_excel_report(results_df, selected_pdf, excel_filename)
                print(f"✅ Excel-rapport lagret i '{excel_filename}'")
            except PermissionError:
                print("❌ FEIL: Kunne ikke lagre Excel-fil. Sjekk at filen ikke er åpen i Excel.")
                print("💡 Lukk filen og prøv igjen, eller gi filen et nytt navn.")
            except Exception as e:
                print(f"❌ FEIL ved Excel-generering: {e}")
            
            print(f"\n📁 Fil opprettet:")
            print(f"   📊 Excel: {excel_filename}")
            
        except KeyboardInterrupt:
            print("\n🛑 Evaluering avbrutt av bruker.")
        except FileNotFoundError as e:
            print(f"\n{e}")
            print("💡 LØSNING: Legg til en PDF-fil med klyngesøknaden i samme mappe som scriptet.")
        except Exception as e:
            if "❌ FEIL:" in str(e):
                print(f"\n{e}")
            else:
                print(f"\n❌ UVENTET FEIL: {e}")
                print("💡 Prøv å kjøre programmet på nytt. Hvis problemet vedvarer, sjekk:")
                print("   - At internettforbindelsen fungerer")
                print("   - At OpenAI API-nøkkelen er gyldig")
                print("   - At PDF-filen ikke er korrupt")
        return
    
    # Innovasjon Norge evaluering
    if valg == "1":
        evaluation_questions = EVALUATION_QUESTIONS_OPPSTART_1
        oppstartstype = "Oppstart 1"
    elif valg == "2":
        evaluation_questions = EVALUATION_QUESTIONS
        oppstartstype = "Oppstart 2"
    else:
        evaluation_questions = EVALUATION_QUESTIONS
        oppstartstype = "Oppstart 3"
    
    # Check if API key is set
    if not openai_api_key:
        print("❌ FEIL: OPENAI_API_KEY ikke funnet.")
        print("💡 LØSNING: Opprett en .env fil i samme mappe med innholdet:")
        print("   OPENAI_API_KEY=din-api-nøkkel-her")
        print("📖 Du kan få en API-nøkkel fra: https://platform.openai.com/api-keys")
        return
    
    try:
        # Read application text
        print("\n🔍 Leser søknadstekst fra PDF...")
        application_text, selected_pdf = read_application_text()
        print(f"✅ PDF lest inn - {len(application_text):,} tegn funnet")
        
        # Create unique filename based on PDF name
        pdf_base_name = selected_pdf.replace('.pdf', '').replace(' ', '_')
        # Remove special characters that might cause issues
        pdf_base_name = re.sub(r'[^\w\-_]', '', pdf_base_name)
        csv_filename = f"evaluering_resultat_{pdf_base_name}.csv"
        excel_filename = f"evaluering_resultat_{pdf_base_name}.xlsx"
        
        # Validate text length
        if len(application_text) < 100:
            print("⚠️  ADVARSEL: Søknadsteksten virker veldig kort. Sjekk at PDF-en inneholder tekst.")
            response = input("Vil du fortsette likevel? (j/n): ")
            if response.lower() != 'j':
                print("🛑 Evaluering avbrutt.")
                return
        
        # Evaluate application
        print("\n🤖 Starter AI-evaluering av søknaden...")
        print("Dette kan ta noen minutter avhengig av søknadens lengde.")
        print("💡 Tips: Du kan avbryte med Ctrl+C hvis nødvendig.")
        
        results_df = evaluate_application(application_text, selected_pdf, evaluation_questions)
        
        # Save results to CSV
        print(f"\n💾 Lagrer resultater til CSV-fil: {csv_filename}")
        try:
            results_df.to_csv(csv_filename, index=False, encoding='utf-8-sig')
            print(f"✅ Resultater lagret i '{csv_filename}'")
        except PermissionError:
            print("❌ FEIL: Kunne ikke lagre CSV-fil. Sjekk at filen ikke er åpen i Excel.")
            print("💡 Lukk filen og prøv igjen, eller gi filen et nytt navn.")
        
        # Create Excel report
        print(f"\n📊 Lager formatert Excel-rapport: {excel_filename}")
        try:
            create_excel_report(results_df, selected_pdf, excel_filename, oppstartstype)
            print(f"✅ Excel-rapport lagret i '{excel_filename}'")
        except PermissionError:
            print("❌ FEIL: Kunne ikke lagre Excel-fil. Sjekk at filen ikke er åpen i Excel.")
            print("💡 Lukk filen og prøv igjen, eller gi filen et nytt navn.")
        except Exception as e:
            print(f"❌ FEIL ved Excel-generering: {e}")
        
        # Print results
        print("\n📊 EVALUERINGSRESULTATER:")
        print("=" * 80)
        print(results_df.to_string(index=False))
        
        # Print summary
        print("\n📈 SAMMENDRAG PER KATEGORI:")
        print("=" * 40)
        summary = results_df.groupby('Kategori')['Score'].mean().round(2)
        for kategori, score in summary.items():
            emoji = "🟢" if score >= 2.5 else "🟡" if score >= 1.5 else "🔴"
            print(f"{emoji} {kategori}: {score}/3.0")
        
        total_score = summary.mean()
        total_emoji = "🟢" if total_score >= 2.5 else "🟡" if total_score >= 1.5 else "🔴"
        print(f"\n🎯 TOTAL GJENNOMSNITTSSCORE: {total_emoji} {total_score:.2f}/3.0")
        
        # Provide interpretation
        if total_score >= 2.5:
            print("🎉 Utmerket søknad! Høy sannsynlighet for godkjenning.")
        elif total_score >= 2.0:
            print("👍 God søknad med potensial. Noen forbedringer kan styrke den.")
        elif total_score >= 1.5:
            print("⚠️  Søknaden trenger forbedringer i flere områder.")
        else:
            print("🔴 Søknaden har betydelige svakheter som bør adresseres.")
        
        print(f"\n📁 Filer opprettet:")
        print(f"   📄 CSV: {csv_filename}")
        print(f"   📊 Excel: {excel_filename}")

    except KeyboardInterrupt:
        print("\n🛑 Evaluering avbrutt av bruker.")
    except FileNotFoundError as e:
        print(f"\n{e}")
        print("💡 LØSNING: Legg til en PDF-fil med søknaden i samme mappe som scriptet.")
    except Exception as e:
        if "❌ FEIL:" in str(e):
            print(f"\n{e}")
        else:
            print(f"\n❌ UVENTET FEIL: {e}")
            print("💡 Prøv å kjøre programmet på nytt. Hvis problemet vedvarer, sjekk:")
            print("   - At internettforbindelsen fungerer")
            print("   - At OpenAI API-nøkkelen er gyldig")
            print("   - At PDF-filen ikke er korrupt")

if __name__ == "__main__":
    main() 