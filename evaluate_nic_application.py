import openai
import pandas as pd
from typing import List, Dict, Tuple
import os
from dotenv import load_dotenv
import PyPDF2
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import re

# Load environment variables from .env file
load_dotenv()

# Set up OpenAI API key
openai.api_key = os.getenv("OPENAI_API_KEY")

# NIC Cluster Program evaluation criteria with weights
NIC_EVALUATION_CRITERIA = {
    "Bakgrunn for klyngen": {
        "weight": 10,
        "questions": [
            "Beskriver klyngens opprinnelse og hvorfor den ble etablert",
            "Beskriver klyngens målgruppe(r), hvilke utfordringer den adresserer og hvorfor medlemmene ikke klarer å løse disse utfordringene individuelt",
            "Beskriver klyngens egnethet til å løse akkurat disse utfordringene"
        ]
    },
    "Klyngens visjon, misjon og hovedmål": {
        "weight": 15,
        "questions": [
            "Beskriver klyngens visjon (hva vil de gjerne bli)",
            "Beskriver klyngens misjon (hvorfor man er til)",
            "Beskriver klyngens SMARTE mål som er målbare og realistiske",
            "Beskriver klyngens SMARTE mål knyttet opp mot klyngeprogrammets mål og ESG",
            "Hvis relevant: Beskriver klyngens bidrag til oppfyllelse av FNs bærekraftsmål"
        ]
    },
    "Fokusområder, aktiviteter, tjenester og gjennomføringsplan": {
        "weight": 25,
        "questions": [
            "Beskriver klyngens fokusområder",
            "Beskriver hva som er forventet resultatmål (konkrete og kvantifiserbare)",
            "Beskriver hvilke aktiviteter klyngen skal arbeide med og hvordan disse underbygger resultatmålene",
            "Beskriver hvordan klyngens aktiviteter er relevante for klyngens medlemmer",
            "Beskriver en gjennomføringsplan som inkluderer hvem som skal gjøre hva og hvordan"
        ]
    },
    "Fremtidige effekter av klyngens arbeid": {
        "weight": 20,
        "questions": [
            "Beskriver fremtidige effekter av klyngens arbeid (både kort og lang sikt)",
            "Beskriver en kobling opp mot klyngeprogrammets mål",
            "Beskriver fremtidige effekter mot ESG, lønnsomhet og medlemsbedriftenes konkurransekraft",
            "Hvis relevant: Beskriver et potensial i klyngens arbeid for medlemmene (eks økte markedsandeler, gevinster og/eller omstilling)"
        ]
    },
    "Klyngens ressursgrunnlag": {
        "weight": 20,
        "questions": [
            "Beskriver klyngens medlemsmasse og sammensetning",
            "Beskriver klyngemedlemmenes motivasjon og ambisjon for medlemskap i klyngen",
            "Beskriver klyngens interne ressursgrunnlag: klyngeledelse, styre og kompetanse",
            "Beskriver klyngens rolle i forhold til klyngens medlemmer og hvordan man organiserer arbeidet",
            "Beskriver sentrale aktiviteter i klyngen og medlemmenes forpliktelser til disse"
        ]
    },
    "Klyngens rolle": {
        "weight": 10,
        "questions": [
            "Beskriver hvorfor klyngen trengs i sitt marked/område",
            "Beskriver hvilken posisjon klyngen har i dag, og hvilken posisjon den skal ta nasjonalt evt internasjonalt",
            "Beskriver hvordan klyngens arbeid kan bidra til realisering av regionale og nasjonale utviklingsplaner",
            "Beskriver klyngens samarbeidspartnere utenfor klyngen (eks andre klynger eller relevante aktører/miljø)",
            "Beskriver klyngens prosessmetodikk for å identifisere og etablere prosjekter/tjenester for klyngens medlemmer"
        ]
    }
}

def read_application_text(filename: str = None) -> tuple[str, str]:
    """Read the application text from a PDF file."""
    if filename is None:
        # Find PDF files in current directory
        pdf_files = glob.glob("*.pdf")
        if not pdf_files:
            raise FileNotFoundError("❌ FEIL: Ingen PDF-filer funnet i mappen. Legg til en PDF-fil med søknaden.")
        
        if len(pdf_files) == 1:
            filename = pdf_files[0]
            print(f"📄 Bruker PDF-fil: {filename}")
        else:
            # Multiple PDF files found, let user choose
            print(f"\n📁 Fant {len(pdf_files)} PDF-filer i mappen:")
            for i, pdf_file in enumerate(pdf_files, 1):
                print(f"  {i}. {pdf_file}")
            
            while True:
                try:
                    choice = input(f"\nVelg hvilken PDF du vil evaluere (1-{len(pdf_files)}): ")
                    choice_num = int(choice)
                    if 1 <= choice_num <= len(pdf_files):
                        filename = pdf_files[choice_num - 1]
                        print(f"📄 Valgt PDF-fil: {filename}")
                        break
                    else:
                        print(f"❌ Ugyldig valg. Velg et tall mellom 1 og {len(pdf_files)}.")
                except ValueError:
                    print("❌ Ugyldig input. Skriv inn et tall.")
                except KeyboardInterrupt:
                    print("\n🛑 Avbrutt av bruker.")
                    raise
    
    try:
        text = ""
        with open(filename, 'rb') as file:
            try:
                pdf_reader = PyPDF2.PdfReader(file)
                total_pages = len(pdf_reader.pages)
                print(f"📖 Leser {total_pages} sider fra PDF...")
                
                for i, page in enumerate(pdf_reader.pages, 1):
                    page_text = page.extract_text()
                    text += page_text + "\n"
                    if i % 5 == 0:  # Show progress every 5 pages
                        print(f"   📄 Behandlet side {i}/{total_pages}")
                
            except PyPDF2.errors.PdfReadError:
                raise Exception(f"❌ FEIL: Kunne ikke lese PDF-filen '{filename}'. Filen kan være korrupt eller passordbeskyttet.")
            except Exception as e:
                raise Exception(f"❌ FEIL: Problem ved lesing av PDF-innhold: {e}")
        
        if not text.strip():
            raise ValueError(f"❌ FEIL: Ingen tekst kunne ekstraheres fra PDF-filen '{filename}'. Filen kan være tom eller inneholde kun bilder.")
        
        print(f"✅ Tekst ekstrahert fra {total_pages} sider")
        return text, filename
        
    except FileNotFoundError:
        raise FileNotFoundError(f"❌ FEIL: PDF-filen '{filename}' ble ikke funnet. Sjekk at filen eksisterer i mappen.")
    except PermissionError:
        raise Exception(f"❌ FEIL: Ingen tilgang til å lese filen '{filename}'. Sjekk filtillatelser.")
    except Exception as e:
        if "❌ FEIL:" in str(e):
            raise  # Re-raise our custom errors as-is
        else:
            raise Exception(f"❌ FEIL: Uventet problem ved lesing av PDF: {e}")

def get_score_from_openai(question: str, application_text: str, category: str) -> Tuple[int, str]:
    """Get score and comment from OpenAI API for a specific question using 0-4 scale."""
    
    scoring_guide = """
    0 = Ikke besvart/vesentlige mangler
    1 = Utydelig/svake beskrivelser og eksempler  
    2 = Mindre gode beskrivelser med vage/overordnede eksempler
    3 = Gode beskrivelser som formidler relevante og konkrete eksempler
    4 = Meget gode beskrivelser som formidler veldig relevante og konkrete eksempler
    """
    
    prompt = f"""Du er en objektiv og konstruktiv ekspert på å evaluere klyngesøknader til NIC (Norwegian Innovation Clusters). 
    
    Evaluer følgende spørsmål for kategorien "{category}":
    {question}
    
    Bruk denne scoringsskalaen:
    {scoring_guide}
    
    Søknadstekst: {application_text}
    
    Vær direkte, objektiv og konstruktiv i din vurdering. Fokuser på å nå målet med evalueringen.
    
    Svar i følgende format:
    Score: [0-4]
    Kommentar: [kort, konstruktiv kommentar]"""
    
    try:
        response = openai.ChatCompletion.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": "Du er en objektiv ekspert på å evaluere klyngesøknader til NIC. Gi konstruktive og direkte vurderinger basert på 0-4 skala."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.2,
            max_tokens=200
        )
        
        # Parse the response
        response_text = response.choices[0].message.content
        
        # Find score and comment lines
        score_lines = [line for line in response_text.split('\n') if line.startswith('Score:')]
        comment_lines = [line for line in response_text.split('\n') if line.startswith('Kommentar:')]
        
        if not score_lines:
            raise ValueError(f"Kunne ikke finne 'Score:' i OpenAI-responsen: {response_text}")
        if not comment_lines:
            raise ValueError(f"Kunne ikke finne 'Kommentar:' i OpenAI-responsen: {response_text}")
        
        score_text = score_lines[0].split(':')[1].strip()
        comment = comment_lines[0].split(':')[1].strip()
        
        # Validate and convert score
        try:
            score = int(score_text)
            if score < 0 or score > 4:
                raise ValueError(f"Score må være mellom 0-4, fikk: {score}")
        except ValueError as ve:
            raise ValueError(f"Kunne ikke konvertere score til tall: '{score_text}'. {ve}")
        
        return score, comment
    
    except openai.error.AuthenticationError:
        raise Exception("❌ FEIL: OpenAI API-nøkkel er ugyldig. Sjekk at OPENAI_API_KEY er riktig satt i .env filen.")
    except openai.error.RateLimitError:
        raise Exception("❌ FEIL: OpenAI API rate limit nådd. Vent litt og prøv igjen.")
    except openai.error.APIConnectionError:
        raise Exception("❌ FEIL: Kunne ikke koble til OpenAI API. Sjekk internettforbindelsen din.")
    except openai.error.InvalidRequestError as e:
        raise Exception(f"❌ FEIL: Ugyldig forespørsel til OpenAI API: {e}")
    except ValueError as ve:
        raise Exception(f"❌ FEIL: Problem med å tolke OpenAI-respons: {ve}")
    except Exception as e:
        raise Exception(f"❌ FEIL: Uventet feil ved OpenAI API-kall: {e}")

def evaluate_nic_application(application_text: str, pdf_filename: str = None) -> pd.DataFrame:
    """Evaluate the NIC cluster application using OpenAI API and return results as DataFrame."""
    results = []
    
    # Calculate total number of questions for progress tracking
    total_questions = sum(len(criteria["questions"]) for criteria in NIC_EVALUATION_CRITERIA.values())
    current_question = 0
    
    for category, criteria in NIC_EVALUATION_CRITERIA.items():
        weight = criteria["weight"]
        questions = criteria["questions"]
        
        print(f"\n📋 Evaluerer kategori: {category} (Vekt: {weight}%)")
        
        for question in questions:
            current_question += 1
            print(f"  ⏳ Spørsmål {current_question}/{total_questions}: {question[:50]}...")
            
            try:
                score, comment = get_score_from_openai(question, application_text, category)
                print(f"  ✅ Score: {score}/4")
                
                results.append({
                    "Kategori": category,
                    "Vekt (%)": weight,
                    "Spørsmål": question,
                    "Score": score,
                    "Kommentar": comment
                })
            except Exception as e:
                print(f"  ❌ Feil ved evaluering av spørsmål: {e}")
                # Add a fallback entry with error information
                results.append({
                    "Kategori": category,
                    "Vekt (%)": weight,
                    "Spørsmål": question,
                    "Score": 0,
                    "Kommentar": f"Feil ved evaluering: {str(e)[:100]}..."
                })
                # Ask user if they want to continue
                print(f"  ⚠️  Vil du fortsette med neste spørsmål? (Trykk Enter for å fortsette, Ctrl+C for å avbryte)")
                try:
                    input()
                except KeyboardInterrupt:
                    print("\n🛑 Evaluering avbrutt av bruker.")
                    raise
    
    print(f"\n🎉 Evaluering fullført! {total_questions} spørsmål behandlet.")
    return pd.DataFrame(results)

def create_nic_excel_report(results_df: pd.DataFrame, pdf_filename: str, excel_filename: str) -> None:
    """Create a formatted Excel report for NIC cluster evaluation."""
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "NIC Klyngeevaluering"
    
    # Define styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    summary_font = Font(bold=True, size=12)
    summary_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    category_font = Font(bold=True, size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Calculate weighted scores by category
    category_scores = {}
    weighted_total = 0
    total_weight = 0
    
    for category in NIC_EVALUATION_CRITERIA.keys():
        category_data = results_df[results_df['Kategori'] == category]
        avg_score = category_data['Score'].mean()
        weight = category_data['Vekt (%)'].iloc[0]
        weighted_score = (avg_score / 4) * weight  # Convert to percentage and apply weight
        
        category_scores[category] = {
            'avg_score': avg_score,
            'weight': weight,
            'weighted_score': weighted_score
        }
        
        weighted_total += weighted_score
        total_weight += weight
    
    # Overall weighted score (out of 100)
    overall_score = weighted_total
    
    # Determine assessment
    if overall_score >= 80:
        assessment = "🎉 Utmerket klyngesøknad! Høy sannsynlighet for godkjenning."
        assessment_color = "C6EFCE"  # Light green
    elif overall_score >= 65:
        assessment = "👍 God klyngesøknad med potensial. Noen forbedringer kan styrke den."
        assessment_color = "FFEB9C"  # Light yellow
    elif overall_score >= 50:
        assessment = "⚠️ Klyngesøknaden trenger forbedringer i flere områder."
        assessment_color = "FFEB9C"  # Light yellow
    else:
        assessment = "🔴 Klyngesøknaden har betydelige svakheter som bør adresseres."
        assessment_color = "FFC7CE"  # Light red
    
    current_row = 1
    
    # Title
    ws.merge_cells(f'A{current_row}:E{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = "NIC KLYNGEEVALUERING"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = header_fill
    title_cell.alignment = center_alignment
    current_row += 2
    
    # PDF filename
    ws[f'A{current_row}'] = "Evaluert søknad:"
    ws[f'A{current_row}'].font = summary_font
    ws[f'B{current_row}'] = pdf_filename
    ws.merge_cells(f'B{current_row}:E{current_row}')
    current_row += 2
    
    # Overall weighted score
    ws[f'A{current_row}'] = "TOTAL VEKTET SCORE:"
    ws[f'A{current_row}'].font = Font(bold=True, size=14)
    ws[f'B{current_row}'] = f"{overall_score:.1f}/100"
    ws[f'B{current_row}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws[f'B{current_row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws[f'B{current_row}'].alignment = center_alignment
    current_row += 1
    
    # Assessment
    ws.merge_cells(f'A{current_row}:E{current_row}')
    assessment_cell = ws[f'A{current_row}']
    assessment_cell.value = assessment
    assessment_cell.font = Font(bold=True, size=12)
    assessment_cell.fill = PatternFill(start_color=assessment_color, end_color=assessment_color, fill_type="solid")
    assessment_cell.alignment = center_alignment
    current_row += 2
    
    # Category summary header
    ws[f'A{current_row}'] = "SAMMENDRAG PER KATEGORI"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws.merge_cells(f'A{current_row}:D{current_row}')
    current_row += 1
    
    # Category summary headers
    headers = ['Kategori', 'Vekt (%)', 'Gj.snitt Score', 'Vektet Bidrag']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = category_font
        cell.fill = summary_fill
        cell.alignment = center_alignment
        cell.border = border
    current_row += 1
    
    # Category summary data
    for category, scores in category_scores.items():
        avg_score = scores['avg_score']
        weight = scores['weight']
        weighted_score = scores['weighted_score']
        
        # Color coding based on average score
        if avg_score >= 3.2:  # 80% of 4
            fill_color = "C6EFCE"  # Green
        elif avg_score >= 2.4:  # 60% of 4
            fill_color = "FFEB9C"  # Yellow
        else:
            fill_color = "FFC7CE"  # Red
        
        ws.cell(row=current_row, column=1, value=category).border = border
        ws.cell(row=current_row, column=2, value=f"{weight}%").border = border
        ws.cell(row=current_row, column=2).alignment = center_alignment
        
        score_cell = ws.cell(row=current_row, column=3, value=f"{avg_score:.1f}/4")
        score_cell.border = border
        score_cell.alignment = center_alignment
        score_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        ws.cell(row=current_row, column=4, value=f"{weighted_score:.1f}").border = border
        ws.cell(row=current_row, column=4).alignment = center_alignment
        
        current_row += 1
    
    current_row += 2
    
    # Detailed results header
    ws[f'A{current_row}'] = "DETALJERTE RESULTATER"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws.merge_cells(f'A{current_row}:E{current_row}')
    current_row += 1
    
    # Column headers for detailed results
    detail_headers = ['Kategori', 'Vekt (%)', 'Spørsmål', 'Score', 'Kommentar']
    for col, header in enumerate(detail_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = category_font
        cell.fill = summary_fill
        cell.alignment = center_alignment
        cell.border = border
    current_row += 1
    
    # Add detailed results
    for _, row in results_df.iterrows():
        ws.cell(row=current_row, column=1, value=row['Kategori']).border = border
        ws.cell(row=current_row, column=2, value=f"{row['Vekt (%)']}%").border = border
        ws.cell(row=current_row, column=2).alignment = center_alignment
        ws.cell(row=current_row, column=3, value=row['Spørsmål']).border = border
        
        score_cell = ws.cell(row=current_row, column=4, value=f"{row['Score']}/4")
        score_cell.border = border
        score_cell.alignment = center_alignment
        
        # Color code scores
        if row['Score'] >= 3.2:  # 80% of 4
            score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif row['Score'] >= 2.4:  # 60% of 4
            score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        comment_cell = ws.cell(row=current_row, column=5, value=row['Kommentar'])
        comment_cell.border = border
        comment_cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        current_row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 80
    
    # Set row heights for better readability
    for row in range(1, current_row):
        if row > len(category_scores) + 15:  # Detailed results section
            ws.row_dimensions[row].height = 60
    
    # Save the workbook
    wb.save(excel_filename)

def main():
    print("🚀 Starter NIC Klyngeevaluering...")
    
    # Check if API key is set
    if not openai.api_key:
        print("❌ FEIL: OPENAI_API_KEY ikke funnet.")
        print("💡 LØSNING: Opprett en .env fil i samme mappe med innholdet:")
        print("   OPENAI_API_KEY=din-api-nøkkel-her")
        print("📖 Du kan få en API-nøkkel fra: https://platform.openai.com/api-keys")
        return
    
    try:
        # Read application text
        print("\n🔍 Leser klyngesøknad fra PDF...")
        application_text, selected_pdf = read_application_text()
        print(f"✅ PDF lest inn - {len(application_text):,} tegn funnet")
        
        # Create unique filename based on PDF name
        pdf_base_name = selected_pdf.replace('.pdf', '').replace(' ', '_')
        # Remove special characters that might cause issues
        pdf_base_name = re.sub(r'[^\w\-_]', '', pdf_base_name)
        excel_filename = f"nic_evaluering_resultat_{pdf_base_name}.xlsx"
        
        # Validate text length
        if len(application_text) < 100:
            print("⚠️  ADVARSEL: Søknadsteksten virker veldig kort. Sjekk at PDF-en inneholder tekst.")
            response = input("Vil du fortsette likevel? (j/n): ")
            if response.lower() != 'j':
                print("🛑 Evaluering avbrutt.")
                return
        
        # Evaluate application
        print("\n🤖 Starter AI-evaluering av klyngesøknaden...")
        print("Dette kan ta noen minutter avhengig av søknadens lengde.")
        print("💡 Tips: Du kan avbryte med Ctrl+C hvis nødvendig.")
        
        results_df = evaluate_nic_application(application_text, selected_pdf)
        
        # Create Excel report
        print(f"\n📊 Lager formatert Excel-rapport: {excel_filename}")
        try:
            create_nic_excel_report(results_df, selected_pdf, excel_filename)
            print(f"✅ Excel-rapport lagret i '{excel_filename}'")
        except PermissionError:
            print("❌ FEIL: Kunne ikke lagre Excel-fil. Sjekk at filen ikke er åpen i Excel.")
            print("💡 Lukk filen og prøv igjen, eller gi filen et nytt navn.")
        except Exception as e:
            print(f"❌ FEIL ved Excel-generering: {e}")
        
        # Calculate and print summary
        print("\n📈 SAMMENDRAG PER KATEGORI:")
        print("=" * 60)
        
        weighted_total = 0
        for category in NIC_EVALUATION_CRITERIA.keys():
            category_data = results_df[results_df['Kategori'] == category]
            avg_score = category_data['Score'].mean()
            weight = category_data['Vekt (%)'].iloc[0]
            weighted_score = (avg_score / 4) * weight
            weighted_total += weighted_score
            
            if avg_score >= 3.2:
                emoji = "🟢"
            elif avg_score >= 2.4:
                emoji = "🟡"
            else:
                emoji = "🔴"
            
            print(f"{emoji} {category}: {avg_score:.1f}/4 (Vekt: {weight}%, Bidrag: {weighted_score:.1f})")
        
        # Overall assessment
        overall_emoji = "🟢" if weighted_total >= 80 else "🟡" if weighted_total >= 50 else "🔴"
        print(f"\n🎯 TOTAL VEKTET SCORE: {overall_emoji} {weighted_total:.1f}/100")
        
        # Provide interpretation
        if weighted_total >= 80:
            print("🎉 Utmerket klyngesøknad! Høy sannsynlighet for godkjenning.")
        elif weighted_total >= 65:
            print("👍 God klyngesøknad med potensial. Noen forbedringer kan styrke den.")
        elif weighted_total >= 50:
            print("⚠️  Klyngesøknaden trenger forbedringer i flere områder.")
        else:
            print("🔴 Klyngesøknaden har betydelige svakheter som bør adresseres.")
        
        print(f"\n📁 Fil opprettet:")
        print(f"   📊 Excel: {excel_filename}")

    except KeyboardInterrupt:
        print("\n🛑 Evaluering avbrutt av bruker.")
    except FileNotFoundError as e:
        print(f"\n{e}")
        print("💡 LØSNING: Legg til en PDF-fil med klyngesøknaden i samme mappe som scriptet.")
    except Exception as e:
        if "❌ FEIL:" in str(e):
            print(f"\n{e}")
        else:
            print(f"\n❌ UVENTET FEIL: {e}")
            print("💡 Prøv å kjøre programmet på nytt. Hvis problemet vedvarer, sjekk:")
            print("   - At internettforbindelsen fungerer")
            print("   - At OpenAI API-nøkkelen er gyldig")
            print("   - At PDF-filen ikke er korrupt")

if __name__ == "__main__":
    main() 